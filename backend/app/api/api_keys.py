from datetime import datetime
from typing import List, Optional
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
import csv
import io
import json

from app.core.database import get_db
from app.core.security import generate_api_key
from app.api.deps import get_current_user
from app.models import User, ApiKey, DataStore, DataStoreItem, WorkspaceMember, Workspace
from app.schemas import ApiKeyCreate, ApiKeyResponse, DataStoreCreate, DataStoreResponse, DataStoreItemResponse, ExportRequest

router = APIRouter(prefix="/api-keys", tags=["API Keys"])
data_router = APIRouter(prefix="/data", tags=["Data"])


@router.post("", response_model=ApiKeyResponse, status_code=status.HTTP_201_CREATED)
async def create_api_key(
    data: ApiKeyCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    raw_key, key_prefix, key_hash = generate_api_key()

    api_key = ApiKey(
        user_id=current_user.id,
        workspace_id=data.workspace_id,
        name=data.name,
        key_hash=key_hash,
        key_prefix=key_prefix,
        scopes=data.scopes or [],
        expires_at=data.expires_at
    )
    db.add(api_key)
    await db.commit()
    await db.refresh(api_key)

    return ApiKeyResponse(
        **{k: getattr(api_key, k) for k in ApiKeyResponse.model_fields.keys() if k != "key"},
        key=raw_key
    )


@router.get("", response_model=List[ApiKeyResponse])
async def list_api_keys(
    workspace_id: Optional[UUID] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    query = select(ApiKey).where(
        ApiKey.user_id == current_user.id,
        ApiKey.revoked_at.is_(None)
    )

    if workspace_id:
        query = query.where(ApiKey.workspace_id == workspace_id)

    query = query.order_by(ApiKey.created_at.desc())
    result = await db.execute(query)
    keys = result.scalars().all()

    return keys


@router.delete("/{key_id}", status_code=status.HTTP_204_NO_CONTENT)
async def revoke_api_key(
    key_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(ApiKey).where(
            ApiKey.id == key_id,
            ApiKey.user_id == current_user.id
        )
    )
    api_key = result.scalar_one_or_none()

    if not api_key:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="API key not found")

    api_key.revoked_at = datetime.utcnow()
    await db.commit()


@data_router.post("", response_model=DataStoreResponse, status_code=status.HTTP_201_CREATED)
async def create_data_store(
    data: DataStoreCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    store = DataStore(
        project_id=data.project_id,
        job_id=data.job_id,
        name=data.name,
        format=data.format,
        schema=data.schema
    )
    db.add(store)
    await db.commit()
    await db.refresh(store)

    return store


@data_router.get("/{store_id}", response_model=DataStoreResponse)
async def get_data_store(
    store_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(DataStore).where(DataStore.id == store_id)
    )
    store = result.scalar_one_or_none()

    if not store:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data store not found")

    return store


@data_router.get("/{store_id}/items", response_model=List[DataStoreItemResponse])
async def list_data_items(
    store_id: UUID,
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=1000),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(DataStoreItem)
        .where(DataStoreItem.data_store_id == store_id)
        .order_by(DataStoreItem.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    return result.scalars().all()


@data_router.post("/export")
async def export_data(
    request: ExportRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    from fastapi.responses import StreamingResponse

    store_result = await db.execute(
        select(DataStore).where(DataStore.id == request.data_store_id)
    )
    store = store_result.scalar_one_or_none()

    if not store:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data store not found")

    query = select(DataStoreItem).where(DataStoreItem.data_store_id == request.data_store_id)

    if request.limit:
        query = query.limit(request.limit)

    result = await db.execute(query)
    items = result.scalars().all()

    data = [item.data for item in items]

    if request.filters:
        filtered_data = []
        for item in data:
            match = True
            for key, value in request.filters.items():
                if item.get(key) != value:
                    match = False
                    break
            if match:
                filtered_data.append(item)
        data = filtered_data

    if request.fields:
        data = [{k: d.get(k) for k in request.fields if k in d} for d in data]

    if request.format == "json":
        output = io.StringIO()
        json.dump(data, output, indent=2, default=str)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode()),
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename={store.name}.json"
            }
        )

    elif request.format == "csv":
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode()),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={store.name}.csv"
            }
        )

    elif request.format == "excel":
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = store.name

            if data:
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={store.name}.xlsx"
                }
            )
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="Excel export requires openpyxl library"
            )

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Invalid export format"
    )


@data_router.get("/{store_id}/stats")
async def get_data_stats(
    store_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    store_result = await db.execute(
        select(DataStore).where(DataStore.id == store_id)
    )
    store = store_result.scalar_one_or_none()

    if not store:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data store not found")

    count_result = await db.execute(
        select(func.count(DataStoreItem.id))
        .where(DataStoreItem.data_store_id == store_id)
    )
    total_items = count_result.scalar() or 0

    sample_result = await db.execute(
        select(DataStoreItem)
        .where(DataStoreItem.data_store_id == store_id)
        .limit(5)
    )
    sample = sample_result.scalars().all()

    return {
        "total_items": total_items,
        "size_bytes": store.size_bytes,
        "format": store.format,
        "sample_data": [item.data for item in sample]
    }
